import os
import re
import openpyxl
import pandas as pd
import tempfile
import json
from flask import Flask, request, render_template, send_file, redirect
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import time

app = Flask(__name__)  # Initialize a Flask application

def read_file_with_multiple_encodings(file_path, encodings=['utf-8', 'iso-8859-1', 'windows-1252']):
    """
    Reads a file with multiple encoding options.
    
    Args:
        file_path (str): The path to the file to be read.
        encodings (list): A list of encodings to try while reading the file.

    Returns:
        str: The content of the file if read successfully.

    Raises:
        UnicodeDecodeError: If all encoding attempts fail.
    """
    for encoding in encodings:
        try:
            # Try to open and read the file with the specified encoding
            with open(file_path, 'r', encoding=encoding) as f:
                return f.read()  # Return the file content
        except UnicodeDecodeError:
            print(f"Failed to decode with {encoding}, trying next encoding.")  # Print error message for failed encoding
    raise UnicodeDecodeError(f"All tried encodings failed for {file_path}")  # Raise error if all encodings fail

def process_html_files(file_addresses, wb, summary_data):
    """
    Processes HTML files and extracts event log data into a worksheet.

    Args:
        file_addresses (list): List of file paths to process.
        wb (Workbook): The Excel workbook to add data to.
        summary_data (list): List to collect summary information about processed files.
    """
    # Create a new worksheet for the event log data
    ws_event_log = wb.create_sheet(title="Output_event_log")

    # Add column headers to the worksheet
    ws_event_log.append(['IP Device', 'System Name', 'Entry ID', 'Date', 'Time', 'Error Type', 'Error Code', 'TaskName', 'Filename', 'Line', 'Parameter'])

    for index, file_address in enumerate(file_addresses, start=1):
        try:
            # Get the base name of the file without the extension
            file_base_name = os.path.splitext(os.path.basename(file_address))[0]
            # Read the content of the HTML file with appropriate encoding
            html_content = read_file_with_multiple_encodings(file_address)

            # Extract the IP Device from the HTML content using regex
            ip_device_match = re.search(r'IP=(\d+\.\d+\.\d+\.\d+)', html_content)
            ip_device = ip_device_match.group(1) if ip_device_match else "Unknown"  # Default to "Unknown" if not found

            # Extract System Name from the HTML content
            system_name_match = re.search(r'System Name:\s*([^\n\r]*)', html_content)
            system_name = system_name_match.group(1).strip().split("<")[0] if system_name_match else "Unknown"  # Default to "Unknown"

            # Define a regex pattern to extract event details from the HTML content
            pattern1 = re.compile(r'<tr><td>(\d+): <font color="#(?:3366FF|606060|009900)">(\d{2}\.\d{2}\.\d{2})\s*(\d{2}:\d{2}:\d{2}):\s*(\S+)\s*(\S+)\s*(\S+)\s*,\s*(\S+)\s*,\s*(\d+)<br>\s*\.+(\S+)')
            # Find all matching events in the content
            events = re.findall(pattern1, html_content)

            if not events:  # If no events were found, skip to the next file
                print(f"No events found in {file_address}.")
                continue

            # Create a DataFrame to store the event data
            df = pd.DataFrame(events, columns=['Entry ID', 'Date', 'Time', 'Error Type', 'Error Code', 'TaskName', 'Filename', 'Line', 'Parameter'])
            # Insert extracted IP Device and System Name into the DataFrame
            df.insert(0, 'System Name', system_name)
            df.insert(0, 'IP Device', ip_device)
            df_sorted = df.sort_values(by='Entry ID')  # Sort the DataFrame by Entry ID

            # Append sorted rows to the worksheet
            for r in dataframe_to_rows(df_sorted, index=False, header=False):
                ws_event_log.append(r)
            # Append summary data for the processed file
            summary_data.append([index, os.path.basename(file_address), ip_device, system_name])
        except FileNotFoundError:
            print(f"File not found: {file_address}")  # Print error if the file was not found
        except Exception as e:
            print(f"Error processing {file_address}: {e}")  # Print any other errors encountered
        finally:
            os.remove(file_address)  # Remove the file after processing

    # Set auto filter for the worksheet and adjust column widths
    ws_event_log.auto_filter.ref = f"A1:{chr(64 + ws_event_log.max_column)}1"

    for col in ws_event_log.columns:
        max_length = 0  # Initialize maximum length for the column
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            if cell.value and len(str(cell.value)) > max_length:  # Check for the maximum length
                max_length = len(cell.value)
        ws_event_log.column_dimensions[column].width = max_length + 2  # Adjust column width

    # Highlight the header row in yellow
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for cell in ws_event_log[1]:
        cell.fill = yellow_fill

def process_log_files(file_addresses, wb, summary_data):
    """
    Processes log files and extracts system log data into a worksheet.

    Args:
        file_addresses (list): List of file paths to process.
        wb (Workbook): The Excel workbook to add data to.
        summary_data (list): List to collect summary information about processed files.
    """
    # Create a new worksheet for system log data
    ws_syslog = wb.create_sheet(title="Syslog")

    # Add column headers to the worksheet
    ws_syslog.append(['System Name', 'Month', 'Date', 'Timestamp', 'Facility', 'Severity Level', 'Mnemonic', 'Message Text', 'Traceback'])

    for index, file_address in enumerate(file_addresses, start=1):
        try:
            # Extract the system name from the file name
            system_name = os.path.basename(file_address).split('.')[0].split('_')[-1]
            # Read the content of the log file with appropriate encoding
            txt_content = read_file_with_multiple_encodings(file_address)

            # Define a regex pattern to extract log events
            pattern1 = re.compile(r'\**(\w{2,3})\s+(\d{1,2}) (\d{2}:\d{2}:\d{2}\.\d{3}\s*\S*): %(\S+)-(\d)-(\w+): (.+)\s*(?:\s*-Traceback=(.+))?')
            # Find all matching events in the content
            events = re.findall(pattern1, txt_content)

            if not events:  # If no events were found, skip to the next file
                print(f"No events found in {file_address}.")
                continue

            # Create a DataFrame to store the log data
            df = pd.DataFrame(events, columns=['Month', 'Date', 'Timestamp', 'Facility', 'Severity Level', 'Mnemonic', 'Message Text', 'Traceback'])
            # Insert extracted System Name into the DataFrame
            df.insert(0, 'System Name', system_name)
            df_sorted = df.sort_values(by='Timestamp')  # Sort the DataFrame by Timestamp

            # Append sorted rows to the worksheet
            for r in dataframe_to_rows(df_sorted, index=False, header=False):
                ws_syslog.append(r)

            # Append summary data for the processed file
            summary_data.append([index, os.path.basename(file_address), system_name])
        except FileNotFoundError:
            print(f"File not found: {file_address}")  # Print error if the file was not found
        except Exception as e:
            print(f"Error processing {file_address}: {e}")  # Print any other errors encountered
        finally:
            os.remove(file_address)  # Remove the file after processing

    # Set auto filter for the worksheet and adjust column widths
    ws_syslog.auto_filter.ref = f"A1:{chr(64 + ws_syslog.max_column)}1"

    for col in ws_syslog.columns:
        max_length = 0  # Initialize maximum length for the column
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            if cell.value and len(str(cell.value)) > max_length:  # Check for the maximum length
                max_length = len(cell.value)
        ws_syslog.column_dimensions[column].width = max_length + 2  # Adjust column width

    # Highlight the header row in yellow
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for cell in ws_syslog[1]:
        cell.fill = yellow_fill

def extract_link_data(file_path):
    """
    Extracts JSON data from the specified log file.

    Args:
        file_path (str): The path to the log file.

    Returns:
        list: A list of extracted JSON data with added Date and Time fields.
    """
    json_list = []
    with open(file_path, 'r') as file:
        for line in file:
            # Check if the line contains JSON data
            if re.search(r'\{.*\}', line):
                # Extract the date and time from the log line
                date_time_match = re.match(r'INFO - (\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})', line)
                if date_time_match:
                    date_time = date_time_match.group(1)
                    date, time = date_time.split(' ')
                    json_str = line[line.find('{'):]
                    try:
                        # Load the JSON data and add Date and Time
                        data = json.loads(json_str)
                        data['Date'] = date
                        data['Time'] = time
                        json_list.append(data)
                    except json.JSONDecodeError:
                        print(f"Invalid JSON: {json_str}")
    return json_list

def process_link_files(file_paths, wb, summary_data):
    """
    Processes multiple link files and appends the extracted data to a worksheet.

    Args:
        file_paths (list): List of file paths to process.
        wb (Workbook): The Excel workbook to add data to.
        summary_data (list): List to collect summary information about processed files.
    """
    # Create a new worksheet for link status data
    ws = wb.create_sheet(title="link_status_jms")

    all_items = []
    for file_path in file_paths:
        # Extract JSON data from each file
        json_data = extract_link_data(file_path)
        for data in json_data:
            # Pop 'items' from data if it exists
            items = data.pop('items', [])
            for item in items:
                # Update item with data fields
                item.update(data)
                all_items.append(item)

    # Check if any items were collected
    if not all_items:
        print("No JSON data found in files.")
        return
    
    # Create a DataFrame from the collected items
    items_df = pd.DataFrame(all_items)
    
    # Reorder columns to have Date and Time first
    columns_order = ['Date', 'Time'] + [col for col in items_df.columns if col not in ['Date', 'Time']]
    items_df = items_df[columns_order]

    # Write DataFrame headers to the worksheet
    headers = list(items_df.columns)
    ws.append(headers)

    # Write DataFrame rows to the worksheet
    for row in dataframe_to_rows(items_df, index=False, header=False):
        ws.append(row)

    # Adjust column widths based on the longest item in each column
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # Define fill styles for headers
    header_fill_green = PatternFill(start_color='8ED973', end_color='8ED973', fill_type='solid')
    header_fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Set header colors based on their names
    green_headers = {'Date', 'Time', 'node', 'apiVersion', 'messageSendTime', 'messageId', 'sequenceNumber', 'itemCount'}

    for col_num, cell in enumerate(ws[1], 1):
        if cell.value in green_headers:
            cell.fill = header_fill_green
        else:
            cell.fill = header_fill_yellow

    # Enable auto filter for the worksheet
    ws.auto_filter.ref = ws.dimensions

    # Add to summary_data (to keep track of processed files)
    summary_data.append([len(file_paths), os.path.basename(file_path)])

def extract_equipment_data(file_path):
    """
    Extracts JSON data from the specified equipment log file.

    Args:
        file_path (str): The path to the log file.

    Returns:
        list: A list of extracted JSON data with added Date and Time fields.
    """
    json_list = []
    with open(file_path, 'r') as file:
        for line in file:
            # Check if the line contains JSON data
            if re.search(r'\{.*\}', line):
                try:
                    # Extract JSON string from the line
                    json_str = line[line.find('{'):]
                    data = json.loads(json_str)

                    # Check if 'messageSendTime' is in the data
                    if 'messageSendTime' in data:
                        datetime_str = data['messageSendTime']
                        date, time = datetime_str.split('T')  # Split date and time
                        time = time.split('+')[0]  # Remove timezone
                        data['Date'] = date  # Add Date field
                        data['Time'] = time.split('.')[0]  # Add Time field without milliseconds
                        json_list.append(data)
                except json.JSONDecodeError as e:
                    print(f"Invalid JSON: {line.strip()}")  # Print invalid JSON lines
    return json_list

# Flatten nested items to a single dictionary
def flatten_items(items):
    """
    Flattens nested items into a single dictionary.

    Args:
        items (list): List of items to flatten.

    Returns:
        list: A list of flattened items.
    """
    flat_items = []
    for item in items:
        flat_item = item.copy()  # Create a copy to avoid modifying the original
        if 'equipmentItems' in item:
            flat_item['equipmentItems'] = str(item['equipmentItems'])  # Convert nested items to string
        flat_items.append(flat_item)
    return flat_items

def process_equipment_files(file_paths, wb, summary_data):
    """
    Processes multiple equipment files and appends the extracted data to a worksheet.

    Args:
        file_paths (list): List of file paths to process.
        wb (Workbook): The Excel workbook to add data to.
        summary_data (list): List to collect summary information about processed files.
    """
    # Create a new worksheet for equipment status data
    ws = wb.create_sheet(title="equipment_status")

    columns_per_sheet = 1000000  # Change the maximum rows per sheet
    row_count = 0

    def add_data_to_sheet(df):
        """
        Adds DataFrame data to the worksheet.

        Args:
            df (DataFrame): DataFrame containing data to add to the worksheet.
        """
        nonlocal row_count
        if row_count == 0:  # Add headers only for the first sheet
            ws.append(list(df.columns))
        for row in dataframe_to_rows(df, index=False, header=False):
            if row_count >= columns_per_sheet:
                break  # Stop adding if row limit is reached
            ws.append(row)
            row_count += 1
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        header_fill_green = PatternFill(start_color='8ED973', end_color='8ED973', fill_type='solid')
        header_fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        green_headers = {'Date', 'Time', 'node', 'apiVersion', 'messageSendTime', 'messageId', 'sequenceNumber', 'itemCount'}
        
        for col_num, cell in enumerate(ws[1], 1):
            if cell.value in green_headers:
                cell.fill = header_fill_green
            else:
                cell.fill = header_fill_yellow
        
        ws.auto_filter.ref = ws.dimensions

    all_items = []        
    
    total_files_processed = 0
    total_items_processed = 0

    for file_path in file_paths:
        json_data = extract_equipment_data(file_path)
        if not json_data:
            print(f"No JSON data found in file: {file_path}")
        total_files_processed += 1
        for data in json_data:
            items = data.pop('items', [])
            if items:  # Check if items exist
                flat_items = flatten_items(items)  # Flatten nested items
                for item in flat_items:
                    item.update(data)
                    all_items.append(item)
                    total_items_processed += 1

    if not all_items:
        print("No JSON data found in files.")
        return

    items_df = pd.DataFrame(all_items)
    
    # Reorder columns
    columns_order = ['Date', 'Time'] + [col for col in items_df.columns if col not in ['Date', 'Time']]
    items_df = items_df[columns_order]

    # Process data and stop if exceeds max rows
    while not items_df.empty:
        if row_count >= columns_per_sheet:
            # Create a new sheet and reset row_count
            ws = wb.create_sheet(title=f"Sheet{len(wb.worksheets) + 1}")
            row_count = 0

        split_index = min(columns_per_sheet - row_count, len(items_df))
        part_df = items_df.iloc[:split_index]
        items_df = items_df.iloc[split_index:]
        
        add_data_to_sheet(part_df)

    # Add summary data to summary_data list
    summary_data.append([total_files_processed, os.path.basename(file_path)])

@app.route('/', methods=['GET', 'POST'])  # Define the route for the index page
def index():
    if request.method == 'POST':  # Check if the request method is POST
        file_name = request.form['file_name']  # Get the output file name from the form
        input_files = request.files.getlist('input_folder')  # Get the uploaded files
        
        if not input_files:  # Check if no files were uploaded
            return "No files uploaded"  # Return an error message

        # Filter uploaded files based on their extensions
        html_files = [file for file in input_files if file.filename.endswith('.html')]
        log_files = [file for file in input_files if file.filename.endswith('.log') and not file.filename.split('/')[1].startswith('link') and not file.filename.split('/')[1].startswith('equipment')]
        text_files = [file for file in input_files if file.filename.endswith('.txt') and not file.filename.split('/')[1].startswith('link') and not file.filename.split('/')[1].startswith('equipment')]
        link_files = [file for file in input_files if file.filename.split('/')[1].startswith('link') and not file.filename.split('/')[1].startswith('equipment')]
        equipment_files = [file for file in input_files if file.filename.split('/')[1].startswith('equipment')]
	
        temp_dir = tempfile.gettempdir()  # Get the temporary directory for saving uploaded files
        wb = openpyxl.Workbook()  # Create a new Excel workbook
        ws_summary = wb.active  # Get the active worksheet
        ws_summary.title = "Number of files"  # Set the title of the worksheet
        ws_summary.append(['Total Number of Files'])  # Add header for total number of files
        
        summary_data = []  # Initialize a list to hold summary data

        # Process HTML files if any are uploaded
        if html_files:
            ws_summary.append(['No.', 'Filename', 'IP Devices', 'System Names'])  # Add headers for HTML files
            html_file_addresses = []  # List to hold paths of saved HTML files
            for html_file in html_files:
                file_path = os.path.join(temp_dir, secure_filename(html_file.filename))  # Create secure file path
                html_file.save(file_path)  # Save the HTML file
                html_file_addresses.append(file_path)  # Add file path to the list
            process_html_files(html_file_addresses, wb, summary_data)  # Process HTML files

        # Process LOG files if any are uploaded
        if log_files:
            ws_summary.append(['No.', 'Filename', 'System Names'])  # Add headers for LOG files
            log_file_addresses = []  # List to hold paths of saved LOG files
            for log_file in log_files:
                file_path = os.path.join(temp_dir, secure_filename(log_file.filename))  # Create secure file path
                log_file.save(file_path)  # Save the LOG file
                log_file_addresses.append(file_path)  # Add file path to the list
            process_log_files(log_file_addresses, wb, summary_data)  # Process LOG files

        # Process TXT files if any are uploaded
        if text_files:
            ws_summary.append(['No.', 'Filename', 'System Names'])  # Add headers for TXT files
            text_file_addresses = []  # List to hold paths of saved TXT files
            for text_file in text_files:
                file_path = os.path.join(temp_dir, secure_filename(text_file.filename))  # Create secure file path
                text_file.save(file_path)  # Save the TXT file
                text_file_addresses.append(file_path)  # Add file path to the list
            process_log_files(text_file_addresses, wb, summary_data)  # Process TXT files (same as log files)

        # Process LINK files if any are uploaded
        if link_files:
            ws_summary.append(['No.', 'Filename'])  # Add headers for LINK files
            link_file_addresses = []  # List to hold paths of saved LINK files
            for link_file in link_files:
                file_path = os.path.join(temp_dir, secure_filename(link_file.filename))  # Create secure file path
                link_file.save(file_path)  # Save the LINK file
                link_file_addresses.append(file_path)  # Add file path to the list
            process_link_files(link_file_addresses, wb, summary_data)  # Process LINK files
        
        # Process Equipment files if any are uploaded
        if equipment_files:
            ws_summary.append(['No.', 'Filename'])  # Add headers for Equipment files
            equipment_file_addresses = []  # List to hold paths of saved Equipment files
            for equipment_file in equipment_files:
                file_path = os.path.join(temp_dir, secure_filename(equipment_file.filename))  # Create secure file path
                equipment_file.save(file_path)  # Save the Equipment file
                equipment_file_addresses.append(file_path)  # Add file path to the list
            process_equipment_files(equipment_file_addresses, wb, summary_data)  # Process Equipment files

        # Update summary sheet with total number of files processed
        total_files = len(summary_data)  # Count total files
        ws_summary.cell(row=1, column=2, value=total_files)  # Update the count in the summary sheet
        for row in summary_data:  # Append each summary row
            ws_summary.append(row)

        # Adjust column widths based on content
        for col in ws_summary.columns:
            max_length = 0  # Initialize max length variable
            column = col[0].column_letter  # Get the letter of the current column
            for cell in col:  # Iterate through each cell
                if cell.value and len(str(cell.value)) > max_length:  # Check for non-empty cells
                    max_length = len(str(cell.value))  # Update max length
            ws_summary.column_dimensions[column].width = max_length + 2  # Set column width

        # Create fill styles for header cells
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow fill for headers
        light_yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')  # Light yellow fill

        ws_summary['A1'].fill = yellow_fill  # Apply fill to the first header cell
        ws_summary['B1'].fill = light_yellow_fill  # Apply fill to the second header cell
        for cell in ws_summary[2]:  # Apply fill to the second row
            cell.fill = yellow_fill

        output_file_path = os.path.join(temp_dir, f'{secure_filename(file_name)}.xlsx')  # Define output file path

        # Remove the default 'Sheet' if it exists
        if 'Sheet' in wb.sheetnames:
            ws = wb['Sheet']  # Get default sheet
            wb.remove(ws)  # Remove default sheet
            
        wb.save(output_file_path)  # Save the workbook to the output path

        # Send the saved file as a response for download
        response = send_file(output_file_path, as_attachment=True, download_name=f'{secure_filename(file_name)}.xlsx')  # Create response

        return response  # Return the response to initiate download

    return render_template('index.html')  # Render the index HTML template for GET requests

if __name__ == '__main__':
    app.run(debug=True)  # Run the Flask application in debug mode
